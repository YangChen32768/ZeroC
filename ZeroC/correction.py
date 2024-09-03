import os
import time
import json
import re
import ast
import csv
import pprint
import copy
from collections import defaultdict
from typing import List
from concurrent.futures import ThreadPoolExecutor, as_completed
import concurrent.futures
import numpy as np
import pandas as pd
from sklearn.metrics import mutual_info_score
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from pympler import asizeof
from operator import itemgetter

from sentence_transformers import SentenceTransformer
from fast_sentence_transformers import FastSentenceTransformer
from langchain.embeddings import Embeddings
from langchain.prompts import (
    PromptTemplate,
    ChatPromptTemplate,
    HumanMessagePromptTemplate,
    SystemMessagePromptTemplate,
    MessagesPlaceholder,
    FewShotChatMessagePromptTemplate,
)
from langchain_openai import ChatOpenAI, OpenAIEmbeddings
from langchain_core.messages import SystemMessage, BaseMessage
from langchain.memory import ConversationBufferMemory, ConversationBufferWindowMemory
from langchain_core.pydantic_v1 import BaseModel, Field
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.runnables import RunnableLambda, RunnablePassthrough
from langchain_community.document_loaders import CSVLoader
from langchain_community.vectorstores import FAISS, utils as faiss_utils
from langchain_community.retrievers import BM25Retriever
from langchain.retrievers import EnsembleRetriever



class Output(BaseModel):
    # The output parser
    chain_of_thought_for_correction: str = Field(...,
                                                 description="The chain_of_thought that led to the proposed correction")
    correction: dict = Field(...,
                             description="the most probable correction for the dirty value")


def form_examples(examps) -> str:
    # 接受一个examples数据，返回用于prompt的字符串
    # examps是一个数组，其中每个元素都是包含了'input'和'output'的字典
    few_shot_examps_str = ''
    for examp in examps:
        examp_str = 'human: ' + \
                    json.dumps(examp['input']) + \
                    '\n' + \
                    'ai: ' + \
                    json.dumps(examp['output']) + \
                    '\n'
        few_shot_examps_str = few_shot_examps_str + examp_str
    return few_shot_examps_str


def get_folder_name(base_path):
    # 检查并创建基础路径
    os.makedirs(base_path, exist_ok=True)

    # 获取所有以 'run-' 开头的文件夹
    existing_folders = [d for d in os.listdir(base_path) 
                        if os.path.isdir(os.path.join(base_path, d)) and d.startswith('run-')]

    # 找到现有文件夹中最大的编号
    max_run = max((int(d.split('-')[1]) for d in existing_folders), default=0)

    # 创建新的运行文件夹
    next_run_folder = f'run-{max_run + 1}'
    next_run_path = os.path.join(base_path, next_run_folder)
    os.makedirs(next_run_path)

    return next_run_path


def select_repair_candidates(embeddings_matrix: np.ndarray, detection: pd.DataFrame, num_clusters: int) -> list:
    # 筛选出需要修复的行
    mask_rows = detection.sum(axis=1) > 0
    filtered_embeddings = embeddings_matrix[mask_rows]
    filtered_detection = detection[mask_rows]
    original_indices = np.where(mask_rows)[0]

    # 对嵌入矩阵进行掩码并重塑形状
    masked_embeddings = filtered_embeddings * filtered_detection.values[..., np.newaxis]
    reshaped_embeddings = masked_embeddings.reshape((filtered_embeddings.shape[0], -1))

    # 聚类
    kmeans = KMeans(n_clusters=num_clusters, random_state=42)
    clusters = kmeans.fit_predict(reshaped_embeddings)

    # 在每个簇中选择最佳修复候选者
    selected_indices = []
    covered_columns = set()

    for i in range(num_clusters):
        cluster_mask = clusters == i
        if np.any(cluster_mask):
            cluster_detection = filtered_detection.iloc[cluster_mask]
            cluster_original_indices = original_indices[cluster_mask]

            def coverage_score(row):
                new_columns = set(row[row == 1].index) - covered_columns
                new_cells = row[row == 1].sum()  # 计算该行的值为1的cell数量
                return len(new_columns), new_cells

            # 计算每一行的覆盖评分，先按新列数量排序，再按新cell数量排序
            coverage_scores = cluster_detection.apply(coverage_score, axis=1)
            best_idx = coverage_scores.idxmax()

            if coverage_scores[best_idx][0] > 0:  # 如果有新列覆盖
                selected_indices.append(cluster_original_indices[best_idx])
                covered_columns.update(cluster_detection.loc[best_idx][cluster_detection.loc[best_idx] == 1].index)

    return selected_indices


def get_auto_prompt_chain(repair_list, column, retriever, CoE, indices, chains, indicators_filtered):
    # 生成Auto-CoT的prompt
    human_message_small_template = HumanMessagePromptTemplate.from_template(human_auto_cot_small)
    human_message_large_template = HumanMessagePromptTemplate.from_template(human_auto_cot_large)
    human_input = '['
    relevant_rows_list = []
    dirty_tuples = []
    dirty_values = []
    clean_values = []
    filtered_header = [header[i] for i in indices]
    column_indicators = indicators_human_repaired[column].values
    column_indices = indices_dict[column]
    column_sums = indicators_human_repaired.iloc[:, column_indices].values.sum(axis=1)
    for row_idx in repair_list:
        if indicators.loc[row_idx, column] == 1:
            dirty_value = dirty_data.loc[row_idx, column]
            dirty_values.append(dirty_value)
            clean_value = clean_data.loc[row_idx, column]
            clean_values.append(clean_value)
            dirty_row = dirty_data.iloc[:, indices].loc[row_idx].tolist()
            dirty_tuples.append(dirty_row)
            relevant_clean_tuples = ''
            embeddings_row = embeddings_matrix[row_idx]
            embeddings_row_filtered = embeddings_row[indices]
            for i in range(len(embeddings_row_filtered)):
                if indicators_filtered.iloc[row_idx, i] == 1:
                    embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
            embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
            embeddings_row_united = embeddings_row_filtered.flatten()
            relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                             k=30)
            relevant_rows_dict_list = [
                {
                    'page_content': row[0].page_content,
                    'index': idx,
                    'score': round(row[1], 2),
                    'target_column': column_indicators[idx],
                    'sum': column_sums[idx]
                }
                for row in relevant_rows
                for idx in [row[0].metadata['index']]
            ]

            sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                        'sum')
            for row in sorted_relevant_rows_dict_list[:3]:
                relevant_clean_tuples += row['page_content'] + '\n'
            relevant_rows_list.append(relevant_clean_tuples)
            human_input += '{' + human_message_small_template.format(
                Dirty_Tuple=format_row(dirty_row, filtered_header),
                Erroneous_value='{' + f'"{column}": "{dirty_value}"' + '}',
                Relevant_clean_tuples=relevant_clean_tuples,
                Correction=clean_value
            ).content + '},'
    human_input += ']'
    prompt_auto_cot = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(sys_auto_cot),
        HumanMessagePromptTemplate.from_template(human_auto_cot_large),
    ],
        partial_variables={
            "examples": examples_auto_cot_str,
        }
    )
    chain_auto_cot = (prompt_auto_cot | llm_auto_cot)
    # while True:
    #     try:
    # promt=prompt_auto_cot.invoke(
    #     {"human_input": human_input})
    # with open('prompt_AutoCoT.txt','w',encoding='utf-8') as f2:
    #     f2.write(promt.to_string())
    repair_result = chain_auto_cot.invoke(
        {"human_input": human_input}).content
    repair_result = repair_result.replace('\n', '')
    cleaned_text = re.sub(r'```json', '', repair_result)
    repair_result = re.sub(r'```', '', cleaned_text)
    repair_result_list = ast.literal_eval(repair_result)
    # 构建examples
    specific_examples_llm = []
    for idx, result in enumerate(repair_result_list):
        dirty_tuple = dirty_tuples[idx]
        json_string = """
                            {
                                "input": {
                                    "Dirty Data Tuple": null,
                                    "Erroneous Value": null,
                                    "Relevant Clean Tuples": null
                                },
                                "output": {
                                    "chain_of_thought_for_correction": null,
                                    "correction": null
                                }
                            }
                            """
        d = json.loads(json_string)
        d['input']['Dirty Data Tuple'] = format_row(dirty_tuple, filtered_header)
        d['input'][
            'Erroneous Value'] = '{' + f'"{column}": "{dirty_values[idx]}"' + '}'
        d['input']['Relevant Clean Tuples'] = relevant_rows_list[idx],
        d['output']['chain_of_thought_for_correction'] = result['chain_of_thought_for_correction']
        d['output']['correction'] = '{"' + column + '\": ' + '"' + clean_values[idx] + '"}'
        specific_examples_llm.append(d)
        #     break
        # except Exception as e:
        #     print('ChatModel请求错误', e)
        #     continue
    few_shot_specific_str = form_examples(specific_examples_llm)
    prompt = ChatPromptTemplate(messages=[
        SystemMessagePromptTemplate.from_template(sys),
        HumanMessagePromptTemplate.from_template(human),
    ],
        partial_variables={
            "general_examples": general_examples_str,
            "specific_examples": few_shot_specific_str,
            "format_instructions": parser.get_format_instructions()
        }
    )
    chain = (
            prompt
            | llm
            | parser
    )
    chains[column] = chain
    # prompt_dict[column] = prompt
    return few_shot_specific_str



    # vectors_matrix = set_dirty_zeros(vectors_matrix, detection)
    texts = dirty_data.values.tolist()
    # 计算互信息以筛选无关列
    normMI = calc_mi_2(dirty_data, target_column)
    normMI = np.round(normMI, decimals=1)
    print(normMI)
    normMI = np.array(normMI)
    threshold = 0.5
    # 在embeddings,texts和detection中分别筛选无关列
    indices = np.where(normMI >= threshold)[0]
    CoE = normMI[indices]
    embeddings_matrix_col_filtered = vectors_matrix[:, indices, :]
    texts_col_filtered = [[row[col] for col in indices] for row in texts]
    detection_col_filtered = detection.iloc[:, indices]
    # 去除待修复列出错的行,embeddings,texts,detection
    # rows_to_keep = np.where(detection_col_filtered[target_column] == 0)[0]
    # embeddings_matrix_col_row_filtered = embeddings_matrix_col_filtered[rows_to_keep, :, :]
    # texts_col_row_filtered = [row for i, row in enumerate(texts_col_filtered) if i in rows_to_keep]
    # detection_col_row_filtered = detection_col_filtered.iloc[rows_to_keep, :]
    # 合并texts
    formatted_rows = []
    header_col_filtered = detection_col_filtered.columns
    for text_row, detection_row in zip(texts_col_filtered, detection_col_filtered.values.tolist()):
        formatted_rows.append(format_row_2(text_row, header_col_filtered, detection_row))
    # reshape embeddings_martix
    embeddings_matrix_col_row_filtered_reshaped = embeddings_matrix_col_filtered.reshape(
        embeddings_matrix_col_filtered.shape[0], -1)
    paired_data = [(text, vector.flatten().tolist()) for text, vector in
                   zip(formatted_rows, embeddings_matrix_col_row_filtered_reshaped)]
    # 生成ids
    ids = [str(i) for i in range(len(dirty_data_human_repaired))]
    db = FAISS.from_embeddings(text_embeddings=paired_data,
                               embedding=embeddingModel,
                               ids=ids,
                               distance_strategy=DistanceStrategy.DOT_PRODUCT)
    retriever = db
    return retriever, indices, CoE


def build_retriever_3(dirty_data, detection, vectors_matrix, target_column):
    texts = dirty_data.values.tolist()
    # 计算互信息以筛选无关列
    normMI = calc_mi_2(dirty_data, target_column)
    normMI = np.round(normMI, decimals=1)
    print(normMI)
    normMI = np.array(normMI)
    threshold = 0.5
    # 在embeddings,texts和detection中分别筛选无关列
    indices = np.where(normMI >= threshold)[0]
    CoE = normMI[indices]
    embeddings_matrix_col_filtered = vectors_matrix[:, indices, :]
    texts_col_filtered = [[row[col] for col in indices] for row in texts]
    detection_col_filtered = detection.iloc[:, indices]
    # 合并texts
    formatted_rows = []
    header_col_filtered = detection_col_filtered.columns
    for text_row, detection_row in zip(texts_col_filtered, detection_col_filtered.values.tolist()):
        formatted_rows.append(format_row_2(text_row, header_col_filtered, detection_row))
    # reshape embeddings_martix
    embeddings_matrix_col_row_filtered_reshaped = embeddings_matrix_col_filtered.reshape(
        embeddings_matrix_col_filtered.shape[0], -1)
    paired_data = [(text, vector.flatten().tolist()) for text, vector in
                   zip(formatted_rows, embeddings_matrix_col_row_filtered_reshaped)]
    ids = [str(i) for i in range(len(dirty_data_human_repaired))]
    # 生成meta_data
    meta_data = [{'index': i} for i in range(len(dirty_data_human_repaired))]

    db = FAISS.from_embeddings(text_embeddings=paired_data,
                               embedding=embeddingModel,
                               metadatas=meta_data,
                               ids=ids,
                               distance_strategy=DistanceStrategy.DOT_PRODUCT)
    retriever = db
    return retriever, indices, CoE


def update_retriever(column):
    # 利用用户修复的结果更新检索器
    ids = [str(i) for i in repair_list]
    retriever_dict[column].delete(ids)
    embeddings_matrix_only_repaired_col_filtered = embeddings_matrix_only_repaired[:, indices_dict[column], :]
    texts = dirty_data_only_repaired.values.tolist()
    texts_col_filtered = [[row[col] for col in indices_dict[column]] for row in texts]
    header_col_filtered = [header[i] for i in indices_dict[column]]
    formatted_rows = []
    for text_row in texts_col_filtered:
        formatted_rows.append(format_row(text_row, header_col_filtered))
    meta_data = [{'index': i} for i in repair_list]
    paired_data = [(text, vector.flatten().tolist()) for text, vector in
                   zip(formatted_rows, embeddings_matrix_only_repaired_col_filtered)]
    retriever_dict[column].add_embeddings(text_embeddings=paired_data,
                                          metadatas=meta_data,
                                          ids=ids
                                          )


def sort_dicts(dict_list, key1, key2, key3):
    return sorted(dict_list, key=lambda x: (x.get(key1, float('inf')),
                                            x.get(key2, float('inf')),
                                            x.get(key3, float('inf'))))


def repair_tableau():
    # 记录llm生成的specific examples
    sp_examps = []
    time_auto_cot_start = time.time()
    print('开始生成自动思维链...')
    chains = {}
    with ThreadPoolExecutor(max_workers=20) as executor:
        futures = []
        for col_idx, column in enumerate(indicators.columns):
            if indicators[column].sum() > 0 and column == 'state':
                print(column)
                print('为其建立检索器...')
                retriever, indices, CoE = build_retriever_3(dirty_data, indicators,
                                                            embeddings_matrix,
                                                            column,
                                                            )
                print('建立完成')
                retriever_dict[column] = retriever
                indices_dict[column] = indices
                CoE_dict[column] = CoE
                indicators_filtered = indicators.iloc[:, indices]
                future = executor.submit(get_auto_prompt_chain,
                                         repair_list,
                                         column,
                                         retriever,
                                         CoE,
                                         indices,
                                         chains,
                                         indicators_filtered
                                         )
                futures.append(future)
        for future in as_completed(futures):
            few_shot_specific_str = future.result()
            sp_examps.append(few_shot_specific_str)
    with open(os.path.join(output_path, 'specific_examples.txt'), 'w', encoding='utf-8') as f_output:
        for sp_examp in sp_examps:
            f_output.write(sp_examp.__str__() + '\n\n')
    print('自动思维链生成完成')
    time_auto_cot_end = time.time()
    print(time_auto_cot_end - time_auto_cot_start)
    # 更新检索器
    time_update_retriever_start = time.time()
    for col_idx, column in enumerate(indicators_human_repaired.columns):
        if indicators_human_repaired[column].sum() > 0:
            update_retriever(column)
    time_update_retriever_end = time.time()
    print(f"更新检索器耗时：{time_update_retriever_end - time_update_retriever_start}")
    total_time = 0
    retriever_time = 0
    dict_creation_time = 0
    sort_time = 0
    print('开始检索...')
    for col_idx, column in enumerate(dirty_data_human_repaired.columns):
        if indicators_human_repaired[column].sum() > 0:
            retriever = retriever_dict[column]
            indices = indices_dict[column]
            CoE = CoE_dict[column]
            temp = indicators_human_repaired.iloc[:, indices]
            for row_idx in range(len(indicators_human_repaired)):
                if indicators_human_repaired.at[row_idx, column] == 1:
                    # 在外层循环中转换为NumPy数组
                    column_indicators = indicators_human_repaired[column].values
                    column_indices = indices_dict[column]
                    column_sums = indicators_human_repaired.iloc[:, column_indices].values.sum(axis=1)
                    start_time = time.time()

                    relevant_clean_tuples = ''
                    embeddings_row = embeddings_matrix[row_idx]
                    embeddings_row_filtered = embeddings_row[indices]
                    for i in range(len(embeddings_row_filtered)):
                        if temp.iloc[row_idx, i] == 1:
                            embeddings_row_filtered[i] = np.zeros(len(embeddings_row_filtered[0]))
                    embeddings_row_filtered = np.multiply(embeddings_row_filtered, CoE[:, np.newaxis])
                    embeddings_row_united = embeddings_row_filtered.flatten()

                    retriever_start = time.time()
                    relevant_rows = retriever.similarity_search_with_score_by_vector(embedding=embeddings_row_united,
                                                                                     k=30)
                    retriever_time += time.time() - retriever_start

                    dict_start = time.time()
                    # 在内层循环中使用NumPy数组
                    relevant_rows_dict_list = [
                        {
                            'page_content': row[0].page_content,
                            'index': idx,
                            'score': round(row[1], 2),
                            'target_column': column_indicators[idx],
                            'sum': column_sums[idx]
                        }
                        for row in relevant_rows
                        for idx in [row[0].metadata['index']]
                    ]
                    dict_creation_time += time.time() - dict_start

                    sort_start = time.time()
                    sorted_relevant_rows_dict_list = sort_dicts(relevant_rows_dict_list, 'score', 'target_column',
                                                                'sum')
                    sort_time += time.time() - sort_start

                    for row in sorted_relevant_rows_dict_list[:3]:
                        relevant_clean_tuples += row['page_content'] + '\n'
                    if column not in retrieved_tuples:
                        retrieved_tuples[column] = {}
                    retrieved_tuples[column][row_idx] = relevant_clean_tuples

                    total_time += time.time() - start_time
    print('检索完成')
    with ThreadPoolExecutor(max_workers=5000) as executor:
        futures = []
        for col_idx, column in enumerate(indicators_human_repaired.columns):
            # 若列中存在错误
            if indicators_human_repaired[column].sum() > 0:
                # 自己首先使用循环组成Input,再将Input传入prompt
                # 遍历列中每个元素
                chain = chains[column]
                for row_idx in range(len(indicators_human_repaired)):
                    if indicators_human_repaired.at[row_idx, column] == 1:
                        # dirty tuple
                        dirty_tuple = dirty_data_human_repaired.iloc[row_idx]
                        dirty_value = dirty_data_human_repaired.at[row_idx, column]
                        # repair the dirty value
                        future = executor.submit(repair_value,
                                                 dirty_tuple,
                                                 column,
                                                 dirty_value,
                                                 row_idx,
                                                 col_idx,
                                                 chain,
                                                 )
                        futures.append(future)
                        # 等待所有的future完成
                    # del retriever
        for future in as_completed(futures):
            future.result()


def repair_value(dirty_tuple, column, dirty_value, index_row, index_col, chain):
    filtered_tuple = dirty_tuple.iloc[indices_dict[column]]
    filtered_header = [header[i] for i in indices_dict[column]]
    dirty_tuple_filtered_str = format_row(filtered_tuple, filtered_header)
    dirty_value_str = '{' + f'"{column}": "{dirty_value}"' + '}'
    dirty_tuple_json = dirty_tuple.to_dict()
    correction = dirty_value
    result_json = dirty_value_str
    relevant_clean_tuples = retrieved_tuples[column][index_row]
    try_num = 0
    while True:
        try:
            repair_result = chain.invoke({'Dirty_Tuple': dirty_tuple_filtered_str,
                                          'Erroneous_value': dirty_value_str,
                                          'Relevant_clean_tuples': relevant_clean_tuples,
                                          })
            result_json = repair_result
            if result_json['correction'].get(column) is None:
                result_json['correction'][column] = 'null'
            correction = result_json['correction'][column]
            break
        except Exception as e:
            print('ChatModel请求错误', e)
            try_num += 1
            if try_num >= 3:
                break
            continue
    corrections.iloc[index_row, index_col] = str(correction)
    log = {'Index': dirty_tuple_json['index'],
           'Dirty_tuple': format_row(dirty_tuple, header),
           'Dirty_value': dirty_value_str,
           'Relevant_clean_tuples': relevant_clean_tuples,
           'Correction': str(result_json)
           }
    logs.append(log)


def cmp_mark(df_A, df_B):
    # 将不一致元素标为高亮
    df_A.fillna('null', inplace=True)
    df_B.fillna('null', inplace=True)
    # 找出不同的元素
    difference = df_A.ne(df_B)
    # 计算不一致元素的总数
    diff_count = difference.sum().sum()
    # 计算不一致元素的百分比
    total_elements = difference.size
    diff_percent = diff_count / total_elements * 100

    # 新建一个Excel文件
    wb = Workbook()
    ws = wb.active

    # 将B DataFrame的数据填充到工作表中
    for r in dataframe_to_rows(df_B, index=False, header=True):
        ws.append(r)

    # 遍历difference DataFrame，为B中与A不同的单元格添加高亮
    for col in range(difference.shape[1]):
        # 注意：openpyxl 从1开始计数，所以我们需要加2（1表示从第1行开始，另外1用于跳过标题行）
        for row in range(difference.shape[0]):
            if difference.iloc[row, col]:
                cell = ws.cell(row=row + 2, column=col + 1)
                # 设置单元格的填充色为黄色
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 打印不一致元素的个数和百分比
    print(f"不一致元素个数: {diff_count}, 百分比: {diff_percent:.2f}%")
    # 保存为xlsx文件
    filename_with_extension = os.path.basename(dirty_data_path)
    filename_without_extension = os.path.splitext(filename_with_extension)[0]
    save_path = os.path.join(output_path, f'{filename_without_extension}-corrected-marked_{MODEL_NAME}.xlsx')
    wb.save(save_path)


def harmonic_mean(a, b):
    # 调和平均值
    return 2 / (1 / a + 1 / b)


def calc_p_r_f(clean_data, dirty_data, corrected_data):
    # 计算P, R, F
    # corrected_data和dirty_data不一致时，表示模型对该元素进行了修复
    mask_bc = corrected_data != dirty_data
    # corrected_data和clean_data一致时，表示模型对该元素修复成功
    mask_ac = clean_data == corrected_data
    corrected_num = mask_bc.sum().sum()
    final_mask1 = mask_ac & mask_bc
    right_corrected_num1 = final_mask1.sum().sum()
    Precision = right_corrected_num1 / corrected_num
    # clean_data和dirty_data不一致时，表示该元素是错误的
    mask_ab = clean_data != dirty_data
    dirty_num = mask_ab.sum().sum()
    # corrected_data和clean_data一致，表示该错误元素成功修复
    final_mask2 = mask_ab & mask_ac
    right_corrected_num2 = final_mask2.sum().sum()
    Recall = right_corrected_num2 / dirty_num
    F1 = harmonic_mean(Precision, Recall)
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        print(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}')
        f_output.write(f'Precision:{Precision}, Recall:{Recall}, F1-Score:{F1}\n')


def save_print_logs(logs):
    with open(os.path.join(output_path, 'output.txt'), 'w', encoding='utf-8') as f_output:
        for log in logs:
            print(f"Dirty_tuple:\n{log['Dirty_tuple']}")
            print(f"Dirty_value:\n{log['Dirty_value']}")
            print(f"Relevant_clean_tuples:\n{log['Relevant_clean_tuples']}")
            print(f"Correction:\n{log['Correction']}")
            f_output.write(f"Dirty_tuple:\n{log['Dirty_tuple']}\n")
            f_output.write(f"Dirty_value:\n{log['Dirty_value']}\n")
            f_output.write(f"Relevant_clean_tuples:\n{log['Relevant_clean_tuples']}\n")
            f_output.write(f"Correction:\n{log['Correction']}\n")


def process_dirty_data(dirty_data, indicators):
    processed_dirty_data = pd.DataFrame(np.where(indicators == 1, 'null', dirty_data), columns=dirty_data.columns)
    return processed_dirty_data


def calc_mi_2(df, target_column):
    # Make sure the target_column exists in dataframe
    assert target_column in df.columns, f"{target_column} does not exist in csv"
    mutual_info_list = []
    # Process each column in the dataframe
    for column in df.columns:
        if column == target_column:
            mutual_info_list.append(mutual_info_score(df[target_column], df[column]))
        else:
            # if column != target_column:
            # Group data by the target column and current column
            grp_df = df.groupby([target_column, column]).filter(lambda x: len(x) > 1)
            # Compute mutual info score
            if not grp_df.empty:
                mutual_info = mutual_info_score(grp_df[target_column], grp_df[column])
                mutual_info_list.append(mutual_info)
            else:
                mutual_info_list.append(0)
    max_mutual_info = max(mutual_info_list) if mutual_info_list else 0
    # Normalize the mutual_info_dict
    normalized_mi = [number / max_mutual_info for number in mutual_info_list]
    # if len(normalized_mi) > 3:
    #     top_three = sorted(normalized_mi, reverse=True)[:3]
    #     normalized_mi = [x if x in top_three else 0 for x in normalized_mi]
    return normalized_mi


def format_row(row, header):
    s = '{' + ', '.join(f'"{col}": "{val}"' for col, val in zip(header, row)) + '}'
    return s


def format_row_2(value, key, detection_row):
    result = {key[i]: value[i] for i in range(len(value)) if detection_row[i] == 0}
    return json.dumps(result)


def set_dirty_zeros(np_array, df):
    # 将脏数据的句向量置为全0作为惩罚
    l = np_array.shape[2]
    for idx, row in df.iterrows():
        for col, value in row.items():
            if value == 1:
                # Set corresponding vector in numpy array to zeros
                np_array[idx, df.columns.get_loc(col)] = np.zeros(l)
    return np_array


class myEmbeddings(Embeddings):
    # 自定义嵌入
    def __init__(self, modelPath):
        self.model = FastSentenceTransformer(modelPath, device="cuda", quantize=True)

    def embed_documents(self, texts: List[str]) -> np.ndarray:
        return self.model.encode(texts, normalize_embeddings=True)

    def embed_query(self, text: str) -> np.ndarray:
        return self.model.encode(text, normalize_embeddings=True)


if __name__ == "__main__":
    human_repair_num = 10
    # param-csv_file_path
    clean_data_path = 'datasets/hospital/hospital_clean.csv'
    dirty_data_path = 'datasets/hospital/hospital_dirty.csv'
    indicators_path = 'datasets/hospital/hospital_dirty_error_detection.csv'
    output_path = get_folder_name('runs_hospital')
    # param-prompts_path
    system_message_path = 'prompt_templates/SystemMessage.txt'
    human_message_path = 'prompt_templates/HumanMessage.txt'
    general_examples_path = 'prompt_templates/examples.txt'
    system_message_auto_cot_path = 'prompt_templates/SystemMessage.txt'
    examples_auto_cot_path = 'prompt_templates/examples_for_AutoCoT.txt'
    human_message_auto_cot_large_path = 'prompt_templates/HumanMessage_for_AutoCoT_large.txt'
    human_message_auto_cot_small_path = 'prompt_templates/HumanMessage_for_AutoCoT_small.txt'
    # param-llm
    MODEL_NAME = 'gpt-3.5-turbo'
    OPENAI_API_BASE = ''
    OPENAI_API_KEY = ''
    TEMPERATURE = 0.5
    MODEL_NAME_auto_cot = 'gpt-4o'
    OPENAI_API_BASE_auto_cot = ''
    OPENAI_API_KEY_auto_cot = ''
    TEMPERATURE_auto_cot = 0
    # param-embedding
    EMBEDDING_MODEL_PATH = r".\LLM-FT-Test\all-MiniLM-L6-v2"
    # 读取examples
    with open(general_examples_path, 'r', encoding='utf-8') as file:
        # 读取文件全部内容
        content = file.read()
        # 使用ast.literal_eval将字符串转换为列表
        general_examples = ast.literal_eval(content)
    general_examples_str = form_examples(general_examples)
    with open(examples_auto_cot_path, 'r', encoding='utf-8') as file:
        content = file.read()
        examples_auto_cot = ast.literal_eval(content)
    examples_auto_cot_str = form_examples(examples_auto_cot)
    parser = JsonOutputParser(pydantic_object=Output)
    # 读取系统提示和用户输入模板
    with open(system_message_path, 'r', encoding='utf-8') as file_1, open(human_message_path, 'r',
                                                                          encoding='utf-8') as file_2:
        sys = file_1.read()
        human = file_2.read()
    with open(system_message_auto_cot_path, 'r', encoding='utf-8') as file_1, open(human_message_auto_cot_large_path,
                                                                                   'r',
                                                                                   encoding='utf-8') as file_2, open(
        human_message_auto_cot_small_path, 'r', encoding='utf-8') as file_3:
        sys_auto_cot = file_1.read()
        human_auto_cot_large = file_2.read()
        human_auto_cot_small = file_3.read()
    llm = ChatOpenAI(
        model_name=MODEL_NAME,
        openai_api_base=OPENAI_API_BASE,
        openai_api_key=OPENAI_API_KEY,
        temperature=TEMPERATURE
    )
    llm_auto_cot = ChatOpenAI(
        model_name=MODEL_NAME_auto_cot,
        openai_api_base=OPENAI_API_BASE_auto_cot,
        openai_api_key=OPENAI_API_KEY_auto_cot,
        temperature=TEMPERATURE_auto_cot
    )
    clean_data = pd.read_csv(clean_data_path, dtype=str, encoding='utf-8', keep_default_na=False, na_values=['', ])
    clean_data.fillna('null', inplace=True)
    dirty_data = pd.read_csv(dirty_data_path, dtype=str, encoding='utf-8', keep_default_na=False, na_values=['', ])
    dirty_data.fillna('null', inplace=True)
    row_count, column_count = dirty_data.shape
    header = dirty_data.columns.tolist()
    indicators = pd.read_csv(indicators_path)
    print(indicators.sum().sum())
    print('开始进行嵌入')
    embeddingModel = myEmbeddings(EMBEDDING_MODEL_PATH)
    logs = []
    corrections = dirty_data.copy()
    start_time = time.time()
    elements_list = dirty_data.values.flatten().tolist()
    embeddings = embeddingModel.embed_documents(elements_list)
    embeddings_matrix = embeddings.reshape(row_count, column_count, len(embeddings[0]))
    print('嵌入完成')
    repair_list = select_repair_candidates(embeddings_matrix, indicators, human_repair_num)
    print('选择完成')
    print(repair_list)
    dirty_data_human_repaired = dirty_data.copy()
    dirty_data_human_repaired.iloc[repair_list] = clean_data.iloc[repair_list]
    indicators_human_repaired = indicators.copy()
    indicators_human_repaired.iloc[repair_list] = 0
    corrections.iloc[repair_list] = clean_data.iloc[repair_list]
    dirty_data_only_repaired = clean_data.iloc[repair_list]
    elements_list_only_repaired = dirty_data_only_repaired.values.flatten().tolist()
    embeddings_only_repaired = embeddingModel.embed_documents(elements_list_only_repaired)
    embeddings_matrix_only_repaired = embeddings_only_repaired.reshape(human_repair_num, column_count,
                                                                       len(embeddings[0]))
    retriever_dict = {}
    indices_dict = {}
    CoE_dict = {}
    retrieved_tuples = {}
    print('开始修复...')
    repair_tableau()
    logs = sorted(logs, key=lambda x: int(x['Index']))
    save_print_logs(logs)
    corrections.to_csv(os.path.join(output_path, 'corrections.csv'), encoding='utf-8')
    cmp_mark(clean_data, corrections)
    calc_p_r_f(clean_data, dirty_data_human_repaired, corrections)
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"程序运行时间为：{execution_time}秒")
    with open(os.path.join(output_path, 'output.txt'), 'a', encoding='utf-8') as f_output:
        f_output.write(f"程序运行时间为：{execution_time}秒")
